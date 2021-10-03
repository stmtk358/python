#! python3
# 機能
#  Google検索結果をExcelに出力する
# 使い方
#  1.Pythonを実行する
# 実行コマンド
#  python google_search_excel.py 検索ワード
#  python google_search_excel.py Python
#  python google_search_excel.py Python 入門

import sys
import time
import openpyxl
import datetime
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import logging

# webdriver_managerのログを抑止する
logging.disable(logging.INFO)

# 検索する
def search(driver, keywords, search_count):
    results = []

    # 検索数まで検索結果を取得する
    while len(results) <= search_count:
        if len(results) == 0:
            # キーワードで検索する
            driver.get("http://google.com/search?q=" + " ".join(keywords))
        else:
            # 次ページがあれば遷移する
            elements = driver.find_elements_by_css_selector("#pnnext")
            if len(elements) == 0:
                break
            elements[0].click()
        time.sleep(3)

        # 検索結果のリンクを取得する
        elements = driver.find_elements_by_css_selector(".g .yuRUbf > a")
        if len(elements) == 0:
            break

        # URLとタイトルを取得する
        for element in elements:
            url = element.get_attribute("href")
            title = element.find_elements_by_css_selector("h3")[0].text
            results.append({"url": url, "title": title})

    return results[:search_count]


# 検索結果をExcelに出力する
def search_excel(keywords, search_count):
    # 検索する
    driver = webdriver.Chrome(ChromeDriverManager().install())
    results = search(driver, keywords, search_count)
    driver.quit()

    if len(results) == 0:
        print("検索結果が見つかりませんでした。")
        return

    # Excelに出力する
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = "キーワード"
    sheet.cell(row=1, column=2).value = " ".join(keywords)
    sheet.cell(row=2, column=1).value = "検索順位"
    sheet.cell(row=2, column=2).value = "タイトル"
    sheet.cell(row=2, column=3).value = "URL"
    for i in range(len(results)):
        sheet.cell(row=3 + i, column=1).value = 1 + i
        sheet.cell(row=3 + i, column=2).value = results[i]["title"]
        sheet.cell(row=3 + i, column=3).value = '=HYPERLINK("{}")'.format(
            results[i]["url"]
        )
    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 60

    sysdate = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    wb.save("検索結果_{}.xlsx".format(sysdate))
    print("検索結果をExcelに出力しました。")


if len(sys.argv) < 2:
    sys.exit("使い方：python google_search_excel.py 検索ワード")

# 検索する（最大20件）
search_excel(sys.argv[1:], 20)
